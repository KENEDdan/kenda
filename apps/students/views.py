import csv
import io
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import (
    ListView, DetailView, CreateView,
    UpdateView, DeleteView, View
)
from django.urls import reverse_lazy
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
from django.db.models import Q

from .models import Student, Grade, Stream, AcademicYear, StudentPromotion
from .forms import StudentEnrolForm, StudentImportForm



@method_decorator(login_required, name='dispatch')
class StudentListView(ListView):
    model = Student
    template_name = 'students/list.html'
    context_object_name = 'students'
    paginate_by = 20

    def get_queryset(self):
        qs = Student.objects.select_related('grade', 'stream', 'academic_year')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(first_name__icontains=q) |
                Q(last_name__icontains=q) |
                Q(student_id__icontains=q) |
                Q(guardian_name__icontains=q)
            )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        grade = self.request.GET.get('grade')
        if grade:
            qs = qs.filter(grade_id=grade)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'All Students'
        ctx['total'] = Student.objects.count()
        ctx['active'] = Student.objects.filter(status='active').count()
        ctx['grades'] = Grade.objects.all()
        ctx['status_choices'] = Student.Status.choices
        return ctx


@method_decorator(login_required, name='dispatch')
class StudentEnrolView(CreateView):
    model = Student
    form_class = StudentEnrolForm
    template_name = 'students/enrol.html'
    success_url = reverse_lazy('students:list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Enrol Student'
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Student enrolled successfully!')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)


@method_decorator(login_required, name='dispatch')
class StudentDetailView(DetailView):
    model = Student
    template_name = 'students/detail.html'
    context_object_name = 'student'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = self.object.get_full_name()
        ctx['status_choices'] = Student.Status.choices
        return ctx


@method_decorator(login_required, name='dispatch')
class StudentEditView(UpdateView):
    model = Student
    form_class = StudentEnrolForm
    template_name = 'students/enrol.html'
    context_object_name = 'student'

    def get_success_url(self):
        return reverse_lazy('students:detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = f"Edit — {self.object.get_full_name()}"
        ctx['is_edit'] = True
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Student record updated.')
        return super().form_valid(form)


@method_decorator(login_required, name='dispatch')
class StudentDeleteView(DeleteView):
    model = Student
    template_name = 'students/confirm_delete.html'
    success_url = reverse_lazy('students:list')
    context_object_name = 'student'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Delete Student'
        return ctx

    def form_valid(self, form):
        messages.success(
            self.request,
            f'Student {self.object.get_full_name()} has been deleted.'
        )
        return super().form_valid(form)


@method_decorator(login_required, name='dispatch')
class StudentStatusChangeView(View):
    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        new_status = request.POST.get('status')
        if new_status in dict(Student.Status.choices):
            old_status = student.get_status_display()
            student.status = new_status
            student.save()
            messages.success(
                request,
                f"{student.get_full_name()}'s status changed from "
                f"{old_status} to {student.get_status_display()}."
            )
        else:
            messages.error(request, 'Invalid status.')
        return redirect('students:detail', pk=pk)


@method_decorator(login_required, name='dispatch')
class StudentIDCardView(DetailView):
    model = Student
    template_name = 'students/id_card.html'
    context_object_name = 'student'

    def get_context_data(self, **kwargs):
        from apps.school.models import SchoolProfile
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = f"ID Card — {self.object.get_full_name()}"
        ctx['school'] = SchoolProfile.get()
        return ctx


@method_decorator(login_required, name='dispatch')
class StudentImportView(View):
    template_name = 'students/import.html'

    def get(self, request):
        return render(request, self.template_name, {
            'form': StudentImportForm(),
            'page_title': 'Import Students',
            'error_rows': [],
        })

    def post(self, request):
        form = StudentImportForm(request.POST, request.FILES)
        error_rows = []

        if not form.is_valid():
            return render(request, self.template_name, {
                'form': form,
                'page_title': 'Import Students',
                'error_rows': error_rows,
            })

        csv_file = request.FILES['csv_file']
        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Please upload a valid .csv file.')
            return render(request, self.template_name, {
                'form': form,
                'page_title': 'Import Students',
                'error_rows': error_rows,
            })

        decoded = csv_file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(decoded))
        success_count = 0

        for i, row in enumerate(reader, start=2):
            try:
                grade = Grade.objects.get(
                    name__iexact=row.get('grade', '').strip()
                )
                academic_year = AcademicYear.objects.get(is_current=True)
                Student.objects.create(
                    first_name=row['first_name'].strip(),
                    last_name=row['last_name'].strip(),
                    middle_name=row.get('middle_name', '').strip(),
                    gender=row.get('gender', 'male').strip().lower(),
                    date_of_birth=row['date_of_birth'].strip(),
                    grade=grade,
                    academic_year=academic_year,
                    guardian_name=row['guardian_name'].strip(),
                    guardian_phone=row['guardian_phone'].strip(),
                    phone=row.get('phone', '').strip(),
                    email=row.get('email', '').strip(),
                    nationality=row.get('nationality', '').strip(),
                )
                success_count += 1
            except Exception as e:
                error_rows.append(f"Row {i}: {str(e)}")

        if success_count:
            messages.success(
                request,
                f'{success_count} student(s) imported successfully!'
            )
        if error_rows:
            messages.warning(
                request,
                f'{len(error_rows)} row(s) failed. Check details below.'
            )

        return render(request, self.template_name, {
            'form': form,
            'page_title': 'Import Students',
            'error_rows': error_rows,
        })


@method_decorator(login_required, name='dispatch')
class StudentPromoteView(View):
    def get(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        return render(request, 'students/promote.html', {
            'student': student,
            'grades': Grade.objects.all(),
            'academic_years': AcademicYear.objects.all(),
            'page_title': f"Promote — {student.get_full_name()}",
            'result_choices': StudentPromotion.Result.choices,
        })

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        to_grade_id = request.POST.get('to_grade')
        to_year_id = request.POST.get('to_academic_year')
        result = request.POST.get('result', 'promoted')
        notes = request.POST.get('notes', '')
        try:
            to_grade = Grade.objects.get(pk=to_grade_id) if to_grade_id else None
            to_year = AcademicYear.objects.get(pk=to_year_id) if to_year_id else None
            StudentPromotion.objects.create(
                student=student,
                from_grade=student.grade,
                to_grade=to_grade,
                from_academic_year=student.academic_year,
                to_academic_year=to_year,
                result=result,
                notes=notes,
                promoted_by=request.user,
            )
            if to_grade:
                student.grade = to_grade
            if to_year:
                student.academic_year = to_year
            if result == 'graduated':
                student.status = 'graduated'
            student.save()
            messages.success(
                request,
                f"{student.get_full_name()} has been "
                f"{'graduated' if result == 'graduated' else 'promoted to ' + (to_grade.name if to_grade else 'next grade')}."
            )
        except Exception as e:
            messages.error(request, f"Promotion failed: {str(e)}")
        return redirect('students:detail', pk=pk)


@method_decorator(login_required, name='dispatch')
class StudentHistoryView(DetailView):
    model = Student
    template_name = 'students/history.html'
    context_object_name = 'student'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = f"History — {self.object.get_full_name()}"
        ctx['promotions'] = self.object.promotions.select_related(
            'from_grade', 'to_grade',
            'from_academic_year', 'to_academic_year',
            'promoted_by'
        )
        return ctx


@method_decorator(login_required, name='dispatch')
class BulkPromotionView(View):
    def get(self, request):
        return render(request, 'students/bulk_promote.html', {
            'grades': Grade.objects.all(),
            'academic_years': AcademicYear.objects.all(),
            'page_title': 'Bulk Promotion',
        })

    def post(self, request):
        from_grade_id = request.POST.get('from_grade')
        to_grade_id = request.POST.get('to_grade')
        from_year_id = request.POST.get('from_academic_year')
        to_year_id = request.POST.get('to_academic_year')
        try:
            from_grade = Grade.objects.get(pk=from_grade_id)
            to_grade = Grade.objects.get(pk=to_grade_id)
            from_year = AcademicYear.objects.get(pk=from_year_id)
            to_year = AcademicYear.objects.get(pk=to_year_id)
            students = Student.objects.filter(
                grade=from_grade,
                academic_year=from_year,
                status='active'
            )
            count = 0
            for student in students:
                StudentPromotion.objects.create(
                    student=student,
                    from_grade=from_grade,
                    to_grade=to_grade,
                    from_academic_year=from_year,
                    to_academic_year=to_year,
                    result='promoted',
                    promoted_by=request.user,
                )
                student.grade = to_grade
                student.academic_year = to_year
                student.save()
                count += 1
            messages.success(
                request,
                f"{count} student(s) promoted from "
                f"{from_grade.name} to {to_grade.name} successfully!"
            )
        except Exception as e:
            messages.error(request, f"Bulk promotion failed: {str(e)}")
        return render(request, 'students/bulk_promote.html', {
            'grades': Grade.objects.all(),
            'academic_years': AcademicYear.objects.all(),
            'page_title': 'Bulk Promotion',
        })


@method_decorator(login_required, name='dispatch')
class StudentExportView(View):
    def get(self, request):
        from apps.school.models import SchoolProfile
        school = SchoolProfile.get()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"

        # Styles
        header_fill = PatternFill(start_color="2D1B69", end_color="2D1B69", fill_type="solid")
        subheader_fill = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        title_font = Font(name="Calibri", bold=True, color="2D1B69", size=14)
        normal_font = Font(name="Calibri", size=10)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="E8E4F5"),
            right=Side(style="thin", color="E8E4F5"),
            top=Side(style="thin", color="E8E4F5"),
            bottom=Side(style="thin", color="E8E4F5"),
        )

        # Title
        ws.merge_cells("A1:L1")
        ws["A1"] = school.name
        ws["A1"].font = title_font
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 28

        # Subtitle
        ws.merge_cells("A2:L2")
        ws["A2"] = f"Student Export Report — Generated on {datetime.date.today().strftime('%d %B %Y')}"
        ws["A2"].font = Font(name="Calibri", color="6B7280", size=10, italic=True)
        ws["A2"].alignment = center
        ws.row_dimensions[2].height = 18

        # Filter summary
        q = request.GET.get('q', '')
        status = request.GET.get('status', '')
        grade = request.GET.get('grade', '')
        filter_text = "All Students"
        if q:
            filter_text += f" | Search: {q}"
        if status:
            filter_text += f" | Status: {status}"
        if grade:
            try:
                g = Grade.objects.get(pk=grade)
                filter_text += f" | Grade: {g.name}"
            except Grade.DoesNotExist:
                pass
        ws.merge_cells("A3:L3")
        ws["A3"] = filter_text
        ws["A3"].font = Font(name="Calibri", color="9CA3AF", size=9)
        ws["A3"].alignment = center
        ws.row_dimensions[3].height = 14
        ws.row_dimensions[4].height = 6

        # Column headers
        columns = [
            ("Student ID", 14), ("First Name", 16), ("Last Name", 16),
            ("Middle Name", 16), ("Gender", 10), ("Grade", 14),
            ("Stream", 12), ("Academic Year", 15), ("Status", 12),
            ("Guardian Name", 20), ("Guardian Phone", 16), ("Enrolled On", 14),
        ]
        for col_idx, (col_name, col_width) in enumerate(columns, start=1):
            cell = ws.cell(row=5, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_width
        ws.row_dimensions[5].height = 22

        # Data
        qs = Student.objects.select_related(
            'grade', 'stream', 'academic_year'
        ).order_by('grade__order', 'last_name', 'first_name')
        if q:
            qs = qs.filter(
                Q(first_name__icontains=q) |
                Q(last_name__icontains=q) |
                Q(student_id__icontains=q)
            )
        if status:
            qs = qs.filter(status=status)
        if grade:
            qs = qs.filter(grade_id=grade)

        for row_idx, student in enumerate(qs, start=6):
            row_fill = PatternFill(
                start_color="FAFBFF" if row_idx % 2 == 0 else "FFFFFF",
                end_color="FAFBFF" if row_idx % 2 == 0 else "FFFFFF",
                fill_type="solid"
            )
            row_data = [
                student.student_id,
                student.first_name,
                student.last_name,
                student.middle_name or '',
                student.get_gender_display(),
                student.grade.name,
                str(student.stream) if student.stream else '',
                str(student.academic_year),
                student.get_status_display(),
                student.guardian_name,
                student.guardian_phone,
                student.enrollment_date.strftime('%d %b %Y'),
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                cell.alignment = left
                cell.border = thin_border
                cell.fill = row_fill
            ws.row_dimensions[row_idx].height = 18

        # Footer
        total_rows = qs.count()
        footer_row = 6 + total_rows
        ws.merge_cells(f"A{footer_row}:L{footer_row}")
        ws[f"A{footer_row}"] = f"Total: {total_rows} student(s)"
        ws[f"A{footer_row}"].font = Font(name="Calibri", bold=True, color="2D1B69", size=10)
        ws[f"A{footer_row}"].fill = subheader_fill
        ws[f"A{footer_row}"].alignment = left
        ws.row_dimensions[footer_row].height = 20

        ws.freeze_panes = "A6"

        # Response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"kenda_students_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response