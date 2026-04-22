import io
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, View
)
from django.urls import reverse_lazy
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
from django.db.models import Q

from .models import Teacher, Department
from .forms import TeacherForm


@method_decorator(login_required, name='dispatch')
class TeacherListView(ListView):
    model = Teacher
    template_name = 'teachers/list.html'
    context_object_name = 'teachers'
    paginate_by = 20

    def get_queryset(self):
        qs = Teacher.objects.select_related('department')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(first_name__icontains=q) |
                Q(last_name__icontains=q) |
                Q(teacher_id__icontains=q) |
                Q(email__icontains=q)
            )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        department = self.request.GET.get('department')
        if department:
            qs = qs.filter(department_id=department)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'All Teachers'
        ctx['total'] = Teacher.objects.count()
        ctx['active'] = Teacher.objects.filter(status='active').count()
        ctx['departments'] = Department.objects.all()
        ctx['status_choices'] = Teacher.Status.choices
        return ctx


@method_decorator(login_required, name='dispatch')
class TeacherAddView(CreateView):
    model = Teacher
    form_class = TeacherForm
    template_name = 'teachers/form.html'
    success_url = reverse_lazy('teachers:list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Add Teacher'
        return ctx

    def form_valid(self, form):
        teacher = form.save(commit=False)
        teacher.save()
        credentials = self._create_user_account(teacher)
        if credentials:
            self.request.session['new_credentials'] = {
                'type': 'Teacher',
                'name': teacher.get_full_name(),
                'teacher_id': teacher.teacher_id,
                'username': credentials['username'],
                'password': credentials['password'],
                'email': credentials['email'],
            }
            return redirect('teachers:credentials')
        messages.success(
            self.request,
            f'{teacher.get_full_name()} added successfully!'
        )
        return redirect(self.success_url)

    def _create_user_account(self, teacher):
        from apps.accounts.models import CustomUser
        import random
        import string

        if teacher.user:
            return None

        base = f"{teacher.first_name.lower()}.{teacher.last_name.lower()}"
        base = ''.join(c for c in base if c.isalnum() or c == '.')
        username = base
        counter = 1
        while CustomUser.objects.filter(username=username).exists():
            username = f"{base}{counter}"
            counter += 1

        temp_password = ''.join(
            random.choices(string.ascii_letters + string.digits, k=10)
        )

        email = teacher.email if teacher.email \
            else f"{username}@teacher.kenda"

        try:
            user = CustomUser.objects.create_user(
                username=username,
                email=email,
                password=temp_password,
                first_name=teacher.first_name,
                last_name=teacher.last_name,
                role='teacher',
            )
            teacher.user = user
            teacher.save()
            return {
                'username': username,
                'password': temp_password,
                'email': email,
            }
        except Exception as e:
            return None

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)


@method_decorator(login_required, name='dispatch')
class TeacherDetailView(DetailView):
    model = Teacher
    template_name = 'teachers/detail.html'
    context_object_name = 'teacher'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = self.object.get_full_name()
        ctx['status_choices'] = Teacher.Status.choices
        return ctx


@method_decorator(login_required, name='dispatch')
class TeacherEditView(UpdateView):
    model = Teacher
    form_class = TeacherForm
    template_name = 'teachers/form.html'
    context_object_name = 'teacher'

    def get_success_url(self):
        return reverse_lazy('teachers:detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = f"Edit — {self.object.get_full_name()}"
        ctx['is_edit'] = True
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Teacher record updated.')
        return super().form_valid(form)


@method_decorator(login_required, name='dispatch')
class TeacherDeleteView(DeleteView):
    model = Teacher
    template_name = 'teachers/confirm_delete.html'
    success_url = reverse_lazy('teachers:list')
    context_object_name = 'teacher'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Delete Teacher'
        return ctx

    def form_valid(self, form):
        messages.success(
            self.request,
            f'Teacher {self.object.get_full_name()} has been removed.'
        )
        return super().form_valid(form)


@method_decorator(login_required, name='dispatch')
class TeacherStatusChangeView(View):
    def post(self, request, pk):
        teacher = get_object_or_404(Teacher, pk=pk)
        new_status = request.POST.get('status')
        if new_status in dict(Teacher.Status.choices):
            old_status = teacher.get_status_display()
            teacher.status = new_status
            teacher.save()
            messages.success(
                request,
                f"{teacher.get_full_name()}'s status changed from "
                f"{old_status} to {teacher.get_status_display()}."
            )
        else:
            messages.error(request, 'Invalid status.')
        return redirect('teachers:detail', pk=pk)


@method_decorator(login_required, name='dispatch')
class TeacherIDCardView(DetailView):
    model = Teacher
    template_name = 'teachers/id_card.html'
    context_object_name = 'teacher'

    def get_context_data(self, **kwargs):
        from apps.school.models import SchoolProfile
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = f"ID Card — {self.object.get_full_name()}"
        ctx['school'] = SchoolProfile.get()
        return ctx


@method_decorator(login_required, name='dispatch')
class TeacherCredentialsView(View):
    def get(self, request):
        credentials = request.session.pop('new_credentials', None)
        if not credentials:
            return redirect('teachers:list')
        return render(request, 'teachers/credentials.html', {
            'credentials': credentials,
            'page_title': 'Account Created',
        })


@method_decorator(login_required, name='dispatch')
class TeacherExportView(View):
    def get(self, request):
        from apps.school.models import SchoolProfile
        school = SchoolProfile.get()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Teachers"

        header_fill = PatternFill(
            start_color="2D1B69", end_color="2D1B69", fill_type="solid"
        )
        subheader_fill = PatternFill(
            start_color="F5F3FF", end_color="F5F3FF", fill_type="solid"
        )
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        title_font = Font(name="Calibri", bold=True, color="2D1B69", size=14)
        normal_font = Font(name="Calibri", size=10)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="E8E4F5"),
            right=Side(style="thin", color="E8E4F5"),
            top=Side(style="thin", color="E8E4F5"),
            bottom=Side(style="thin", color="E8E4F5"),
        )

        ws.merge_cells("A1:J1")
        ws["A1"] = school.name
        ws["A1"].font = title_font
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:J2")
        ws["A2"] = f"Teacher Export — {datetime.date.today().strftime('%d %B %Y')}"
        ws["A2"].font = Font(
            name="Calibri", color="6B7280", size=10, italic=True
        )
        ws["A2"].alignment = center
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 6

        columns = [
            ("Teacher ID", 14), ("First Name", 16), ("Last Name", 16),
            ("Gender", 10), ("Department", 18), ("Employment Type", 18),
            ("Status", 12), ("Qualification", 22),
            ("Phone", 16), ("Email", 24),
        ]

        for col_idx, (col_name, col_width) in enumerate(columns, start=1):
            cell = ws.cell(row=4, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = col_width
        ws.row_dimensions[4].height = 22

        qs = Teacher.objects.select_related('department').order_by(
            'last_name', 'first_name'
        )
        q = self.request.GET.get('q', '')
        status = self.request.GET.get('status', '')
        department = self.request.GET.get('department', '')
        if q:
            qs = qs.filter(
                Q(first_name__icontains=q) |
                Q(last_name__icontains=q) |
                Q(teacher_id__icontains=q)
            )
        if status:
            qs = qs.filter(status=status)
        if department:
            qs = qs.filter(department_id=department)

        for row_idx, teacher in enumerate(qs, start=5):
            row_fill = PatternFill(
                start_color="FAFBFF" if row_idx % 2 == 0 else "FFFFFF",
                end_color="FAFBFF" if row_idx % 2 == 0 else "FFFFFF",
                fill_type="solid"
            )
            row_data = [
                teacher.teacher_id,
                teacher.first_name,
                teacher.last_name,
                teacher.get_gender_display(),
                str(teacher.department) if teacher.department else '',
                teacher.get_employment_type_display(),
                teacher.get_status_display(),
                teacher.qualification,
                teacher.phone,
                teacher.email,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                cell.alignment = left
                cell.border = thin_border
                cell.fill = row_fill
            ws.row_dimensions[row_idx].height = 18

        total_rows = qs.count()
        footer_row = 5 + total_rows
        ws.merge_cells(f"A{footer_row}:J{footer_row}")
        ws[f"A{footer_row}"] = f"Total: {total_rows} teacher(s)"
        ws[f"A{footer_row}"].font = Font(
            name="Calibri", bold=True, color="2D1B69", size=10
        )
        ws[f"A{footer_row}"].fill = subheader_fill
        ws[f"A{footer_row}"].alignment = left
        ws.row_dimensions[footer_row].height = 20
        ws.freeze_panes = "A5"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"kenda_teachers_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response