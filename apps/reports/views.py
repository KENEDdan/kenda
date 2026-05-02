import io
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView, View
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Count, Sum, Q, Avg

from apps.students.models import Student, Grade, AcademicYear
from apps.teachers.models import Teacher
from apps.attendance.models import AttendanceRecord
from apps.fees.models import Invoice, Payment
from apps.academics.models import ClassRoom


@method_decorator(login_required, name='dispatch')
class ReportsDashboardView(TemplateView):
    template_name = 'reports/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Reports & Analytics'

        # Student stats
        ctx['total_students'] = Student.objects.count()
        ctx['active_students'] = Student.objects.filter(status='active').count()
        ctx['male_students'] = Student.objects.filter(
            status='active', gender='male'
        ).count()
        ctx['female_students'] = Student.objects.filter(
            status='active', gender='female'
        ).count()

        # Teacher stats
        ctx['total_teachers'] = Teacher.objects.count()
        ctx['active_teachers'] = Teacher.objects.filter(status='active').count()

        # Attendance stats
        total_records = AttendanceRecord.objects.count()
        present_records = AttendanceRecord.objects.filter(
            status='present'
        ).count()
        ctx['attendance_rate'] = (
            round(present_records / total_records * 100, 1)
            if total_records > 0 else 0
        )

        # Fee stats
        ctx['total_invoiced'] = Invoice.objects.aggregate(
            t=Sum('total_amount')
        )['t'] or 0
        ctx['total_collected'] = Invoice.objects.aggregate(
            t=Sum('amount_paid')
        )['t'] or 0
        ctx['collection_rate'] = (
            round(ctx['total_collected'] / ctx['total_invoiced'] * 100, 1)
            if ctx['total_invoiced'] > 0 else 0
        )

        # Students per grade
        ctx['students_by_grade'] = Grade.objects.annotate(
            count=Count('students', filter=Q(students__status='active'))
        ).order_by('order')

        return ctx


@method_decorator(login_required, name='dispatch')
class StudentReportView(View):
    template_name = 'reports/students.html'

    def get(self, request):
        grades = Grade.objects.annotate(
            total=Count('students'),
            active=Count(
                'students',
                filter=Q(students__status='active')
            ),
            male=Count(
                'students',
                filter=Q(students__status='active', students__gender='male')
            ),
            female=Count(
                'students',
                filter=Q(students__status='active', students__gender='female')
            ),
        ).order_by('order')

        total_students = Student.objects.count()
        active = Student.objects.filter(status='active').count()
        graduated = Student.objects.filter(status='graduated').count()
        suspended = Student.objects.filter(status='suspended').count()

        return render(request, self.template_name, {
            'grades': grades,
            'total_students': total_students,
            'active': active,
            'graduated': graduated,
            'suspended': suspended,
            'page_title': 'Student Report',
        })


@method_decorator(login_required, name='dispatch')
class AttendanceReportView(View):
    template_name = 'reports/attendance.html'

    def get(self, request):
        classrooms = ClassRoom.objects.filter(
            is_active=True
        ).select_related('grade', 'stream')

        classroom_stats = []
        for classroom in classrooms:
            records = AttendanceRecord.objects.filter(classroom=classroom)
            total = records.count()
            present = records.filter(status='present').count()
            absent = records.filter(status='absent').count()
            late = records.filter(status='late').count()
            rate = round(present / total * 100, 1) if total > 0 else 0
            classroom_stats.append({
                'classroom': classroom,
                'total': total,
                'present': present,
                'absent': absent,
                'late': late,
                'rate': rate,
            })

        return render(request, self.template_name, {
            'classroom_stats': classroom_stats,
            'page_title': 'Attendance Report',
        })


@method_decorator(login_required, name='dispatch')
class FinancialReportView(View):
    template_name = 'reports/fees.html'

    def get(self, request):
        total_invoiced = Invoice.objects.aggregate(
            t=Sum('total_amount')
        )['t'] or 0
        total_collected = Invoice.objects.aggregate(
            t=Sum('amount_paid')
        )['t'] or 0
        total_outstanding = total_invoiced - total_collected

        paid_count = Invoice.objects.filter(status='paid').count()
        partial_count = Invoice.objects.filter(status='partial').count()
        unpaid_count = Invoice.objects.filter(
            status__in=['unpaid', 'overdue']
        ).count()

        # Grade-wise collection
        grade_stats = []
        for grade in Grade.objects.all():
            invoices = Invoice.objects.filter(student__grade=grade)
            invoiced = invoices.aggregate(t=Sum('total_amount'))['t'] or 0
            collected = invoices.aggregate(t=Sum('amount_paid'))['t'] or 0
            rate = round(collected / invoiced * 100, 1) if invoiced > 0 else 0
            grade_stats.append({
                'grade': grade,
                'invoiced': invoiced,
                'collected': collected,
                'outstanding': invoiced - collected,
                'rate': rate,
            })

        # Recent payments
        recent_payments = Payment.objects.select_related(
            'invoice__student'
        ).order_by('-created_at')[:10]

        return render(request, self.template_name, {
            'total_invoiced': total_invoiced,
            'total_collected': total_collected,
            'total_outstanding': total_outstanding,
            'paid_count': paid_count,
            'partial_count': partial_count,
            'unpaid_count': unpaid_count,
            'grade_stats': grade_stats,
            'recent_payments': recent_payments,
            'page_title': 'Financial Report',
        })


@method_decorator(login_required, name='dispatch')
class ExportStudentSummaryView(View):
    def get(self, request):
        from apps.school.models import SchoolProfile
        school = SchoolProfile.get()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Summary"

        header_fill = PatternFill(
            start_color="2D1B69", end_color="2D1B69", fill_type="solid"
        )
        header_font = Font(
            name="Calibri", bold=True, color="FFFFFF", size=11
        )
        title_font = Font(
            name="Calibri", bold=True, color="2D1B69", size=14
        )
        normal_font = Font(name="Calibri", size=10)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        thin = Border(
            left=Side(style="thin", color="E8E4F5"),
            right=Side(style="thin", color="E8E4F5"),
            top=Side(style="thin", color="E8E4F5"),
            bottom=Side(style="thin", color="E8E4F5"),
        )

        ws.merge_cells("A1:F1")
        ws["A1"] = f"{school.name} — Student Summary Report"
        ws["A1"].font = title_font
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:F2")
        ws["A2"] = f"Generated: {datetime.date.today().strftime('%d %B %Y')}"
        ws["A2"].font = Font(
            name="Calibri", color="6B7280", size=10, italic=True
        )
        ws["A2"].alignment = center
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 6

        cols = [
            ("Grade", 20), ("Total Students", 18),
            ("Active", 12), ("Male", 10),
            ("Female", 10), ("Graduated", 12),
        ]
        for i, (name, width) in enumerate(cols, start=1):
            cell = ws.cell(row=4, column=i, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = width
        ws.row_dimensions[4].height = 22

        grades = Grade.objects.annotate(
            total=Count('students'),
            active=Count(
                'students', filter=Q(students__status='active')
            ),
            male=Count(
                'students',
                filter=Q(students__status='active', students__gender='male')
            ),
            female=Count(
                'students',
                filter=Q(students__status='active', students__gender='female')
            ),
            graduated=Count(
                'students', filter=Q(students__status='graduated')
            ),
        ).order_by('order')

        for i, grade in enumerate(grades, start=5):
            fill = PatternFill(
                start_color="FAFBFF" if i % 2 == 0 else "FFFFFF",
                end_color="FAFBFF" if i % 2 == 0 else "FFFFFF",
                fill_type="solid"
            )
            row = [
                grade.name, grade.total, grade.active,
                grade.male, grade.female, grade.graduated,
            ]
            for j, val in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.font = normal_font
                cell.alignment = left
                cell.border = thin
                cell.fill = fill
            ws.row_dimensions[i].height = 18

        ws.freeze_panes = "A5"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"kenda_student_summary_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response